"""
Write the HS6 coverage report: one table.

For every partner and every year, three counts of HS6 product lines -

    UZ     how many Uzbekistan recorded importing from that partner
    Ptn    how many the partner recorded exporting to Uzbekistan
    Shown  how many BOTH recorded, which is what survives to the dashboard

Shown is the whole of the cleaning rule. There is no size floor, no chapter
exclusion and no requirement about the direction of the gap: a line is comparable
when two books describe it.
"""
import json
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "out")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEST = os.path.join(HERE, "hs6_lines_by_country_year.xlsx")

bc = pd.read_pickle(os.path.join(HERE, "by_country.pkl"))
cells = pd.read_pickle(os.path.join(HERE, "hs6_cells.pkl"))
fstats = json.load(open(os.path.join(HERE, "funnel_stats.json"), encoding="utf-8"))
recon = fstats["recon"]

YEARS = list(range(2017, 2025))
FONT = "Arial"
INK = "1F2933"
HEAD = PatternFill("solid", fgColor="1F3A5F")
YEAR_FILL = PatternFill("solid", fgColor="2C5282")
SHOWN_FILL = PatternFill("solid", fgColor="E8EDF2")
NOTE_FILL = PatternFill("solid", fgColor="FFF8E1")
thin = Side(style="thin", color="BFC9D4")
N0 = "#,##0"
PCT = "0.0%"

# ---- the three counts, per partner and year ------------------------------
cells = cells[(cells["pe"] > 0) | (cells["ui"] > 0)].copy()
cells["uz"] = cells["ui"] > 0
cells["ptn"] = cells["pe"] > 0
cells["both"] = cells["uz"] & cells["ptn"]
g = (cells.groupby(["iso", "year"], observed=True)[["uz", "ptn", "both"]].sum()
     .astype(int).reset_index())
counts = {(r.iso, r.year): (r.uz, r.ptn, r.both) for r in g.itertuples()}

wb = Workbook()
ws = wb.active
ws.title = "HS6 lines by country and year"

# ---- title and legend ----------------------------------------------------
ws["A1"] = "Uzbekistan mirror trade - HS6 product lines by country and year"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
LEGEND = [
    "Each figure counts distinct HS6 product lines for that partner in that year.",
    "UZ = lines Uzbekistan recorded importing.   Ptn = lines the partner recorded exporting to "
    "Uzbekistan.   Shown = lines BOTH recorded, which is what the dashboard can compare.",
    "Shown is the entire cleaning rule: no size floor, no chapter exclusion, no condition on the "
    "direction of the gap. A line is comparable when two books describe it.",
    "Where Shown is far below UZ, the partner did not file those lines - a reporting gap on their "
    "side, which no threshold can recover. Zero Shown means no mirror comparison was possible that "
    "year, not that trade was zero.",
]
for i, line in enumerate(LEGEND):
    c = ws.cell(row=2 + i, column=1, value=line)
    c.font = Font(name=FONT, size=9.5, italic=(i > 0), color="52606D")
    c.alignment = Alignment(vertical="top")

HDR1, HDR2, FIRST = 7, 8, 9

# ---- header: a merged year band over three sub-columns -------------------
ws.merge_cells(start_row=HDR1, start_column=1, end_row=HDR2, end_column=1)
ws.merge_cells(start_row=HDR1, start_column=2, end_row=HDR2, end_column=2)
ws.cell(row=HDR1, column=1, value="ISO")
ws.cell(row=HDR1, column=2, value="Country")
ws.column_dimensions["A"].width = 7
ws.column_dimensions["B"].width = 26

for i, y in enumerate(YEARS):
    base = 3 + 3 * i
    ws.merge_cells(start_row=HDR1, start_column=base, end_row=HDR1, end_column=base + 2)
    ws.cell(row=HDR1, column=base, value=str(y))
    for j, lab in enumerate(("UZ", "Ptn", "Shown")):
        ws.cell(row=HDR2, column=base + j, value=lab)
        ws.column_dimensions[get_column_letter(base + j)].width = 7.5

TOT = 3 + 3 * len(YEARS)
ws.merge_cells(start_row=HDR1, start_column=TOT, end_row=HDR1, end_column=TOT + 2)
ws.cell(row=HDR1, column=TOT, value="All years")
for j, lab in enumerate(("UZ", "Ptn", "Shown")):
    ws.cell(row=HDR2, column=TOT + j, value=lab)
    ws.column_dimensions[get_column_letter(TOT + j)].width = 9
PCT_COL = TOT + 3
ws.merge_cells(start_row=HDR1, start_column=PCT_COL, end_row=HDR2, end_column=PCT_COL)
ws.cell(row=HDR1, column=PCT_COL, value="Shown as % of Uzbekistan's lines")
ws.column_dimensions[get_column_letter(PCT_COL)].width = 13
YRS_COL = PCT_COL + 1
ws.merge_cells(start_row=HDR1, start_column=YRS_COL, end_row=HDR2, end_column=YRS_COL)
ws.cell(row=HDR1, column=YRS_COL, value="Years with any shown")
ws.column_dimensions[get_column_letter(YRS_COL)].width = 12

NCOLS = YRS_COL
for row in (HDR1, HDR2):
    ws.row_dimensions[row].height = 34 if row == HDR1 else 16
    for c in range(1, NCOLS + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10 if row == HDR1 else 9, bold=True, color="FFFFFF")
        cell.fill = HEAD if row == HDR1 else YEAR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---- body ----------------------------------------------------------------
SHOWN_COLS = [3 + 3 * i + 2 for i in range(len(YEARS))]
for j, row in enumerate(bc.itertuples(), start=FIRST):
    ws.cell(row=j, column=1, value=row.iso)
    ws.cell(row=j, column=2, value=row.name)
    for i, y in enumerate(YEARS):
        uz, ptn, both = counts.get((row.iso, y), (0, 0, 0))
        base = 3 + 3 * i
        ws.cell(row=j, column=base, value=uz)
        ws.cell(row=j, column=base + 1, value=ptn)
        ws.cell(row=j, column=base + 2, value=both)
    for k, off in enumerate((0, 1, 2)):
        refs = ",".join(f"{get_column_letter(3 + 3 * i + off)}{j}" for i in range(len(YEARS)))
        ws.cell(row=j, column=TOT + k, value=f"=SUM({refs})")
    uzc, shc = get_column_letter(TOT), get_column_letter(TOT + 2)
    ws.cell(row=j, column=PCT_COL, value=f'=IF({uzc}{j}=0,"",{shc}{j}/{uzc}{j})')
    shown_refs = ",".join(f"{get_column_letter(c)}{j}" for c in SHOWN_COLS)
    ws.cell(row=j, column=YRS_COL, value=f"=COUNTIF({shown_refs},\">0\")")

LAST = FIRST + len(bc) - 1
for r in range(FIRST, LAST + 1):
    for c in range(1, NCOLS + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=9.5, color=INK)
        cell.number_format = PCT if c == PCT_COL else N0 if c > 2 else "General"
    for c in SHOWN_COLS + [TOT + 2]:
        ws.cell(row=r, column=c).fill = SHOWN_FILL

ws.freeze_panes = f"C{FIRST}"
ws.auto_filter.ref = f"A{HDR2}:{get_column_letter(NCOLS)}{LAST}"

# ---- total row -----------------------------------------------------------
TROW = LAST + 1
ws.cell(row=TROW, column=2, value="TOTAL - all partners")
for c in range(3, TOT + 3):
    L = get_column_letter(c)
    ws.cell(row=TROW, column=c, value=f"=SUM({L}{FIRST}:{L}{LAST})")
uzc, shc = get_column_letter(TOT), get_column_letter(TOT + 2)
ws.cell(row=TROW, column=PCT_COL, value=f"={shc}{TROW}/{uzc}{TROW}")
for c in range(1, NCOLS + 1):
    cell = ws.cell(row=TROW, column=c)
    cell.font = Font(name=FONT, size=10, bold=True, color=INK)
    cell.border = Border(top=Side(style="thin", color="1F3A5F"))
    cell.number_format = PCT if c == PCT_COL else N0 if c > 2 else "General"

# ---- notes ---------------------------------------------------------------
NOTES = [
    f"Source: data/uzbekistan_mirror_trade_hs2017_fixed_2017_2024.xlsx, all 16 HS6 sheets "
    f"({recon['workbookCells']:,} partner x product x year cells, 2017-2024). Partner exports are "
    f"FOB, Uzbekistan imports CIF; no freight adjustment is applied to these counts.",
    f"Verification: rebuilt from the workbook independently of the dashboard pipeline and "
    f"reconciled cell by cell against src/data/cells.json - {recon['common']:,} cells matched, "
    f"largest value difference $0 on both sides, nothing missing from the rebuild. The "
    f"{recon['onlyWorkbook']} cells the rebuild has and the shipped data does not are sub-dollar "
    f"rows, $47.32 in total, that round to zero in scripts/extract-excel.py.",
    "A count of lines is not a count of value: a partner can file few lines worth a great deal, or "
    "many worth very little. Rebuild with analysis/hs6_coverage_report.py.",
]
for i, n in enumerate(NOTES):
    r = TROW + 2 + i
    c = ws.cell(row=r, column=1, value=n)
    c.font = Font(name=FONT, size=9, italic=True, color="52606D")
    c.alignment = Alignment(vertical="top")

wb.calculation.fullCalcOnLoad = True
wb.save(DEST)
print("wrote", DEST, round(os.path.getsize(DEST) / 1e6, 2), "MB")
print(f"  {len(bc)} partners x {len(YEARS)} years x 3 counts")
