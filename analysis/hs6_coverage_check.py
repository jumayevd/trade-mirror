"""
Audit the report without a recalculation engine.

LibreOffice is not installed here, so openpyxl's formulas carry no cached values.
Every one is therefore checked by hand: the reference is parsed, evaluated against
the values actually written into the sheet, and compared with counts derived
straight from the workbook extract. A formula pointing at the wrong column, the
wrong row or the wrong range fails here rather than reaching the reader.
"""
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "out")
PATH = os.path.join(HERE, "hs6_lines_by_country_year.xlsx")

YEARS = list(range(2017, 2025))
bc = pd.read_pickle(os.path.join(HERE, "by_country.pkl"))
cells = pd.read_pickle(os.path.join(HERE, "hs6_cells.pkl"))
wb = load_workbook(PATH)
ws = wb["HS6 lines by country and year"]
fails, checks = [], 0


def ok(cond, msg):
    global checks
    checks += 1
    if not cond:
        fails.append(msg)


def val(ref):
    return ws[ref].value


# counts rebuilt here independently of the report script
cells = cells[(cells["pe"] > 0) | (cells["ui"] > 0)].copy()
cells["uz"] = cells["ui"] > 0
cells["ptn"] = cells["pe"] > 0
cells["both"] = cells["uz"] & cells["ptn"]
g = cells.groupby(["iso", "year"], observed=True)[["uz", "ptn", "both"]].sum().astype(int)
want = {(i, y): tuple(r) for (i, y), r in g.iterrows()}

HDR1, HDR2, FIRST = 7, 8, 9
LAST = FIRST + len(bc) - 1
TOT = 3 + 3 * len(YEARS)
PCT_COL, YRS_COL = TOT + 3, TOT + 4

# ---- header shape --------------------------------------------------------
ok(val("A7") == "ISO", "A7 header")
ok(val("B7") == "Country", "B7 header")
for i, y in enumerate(YEARS):
    base = 3 + 3 * i
    ok(ws.cell(row=HDR1, column=base).value == str(y), f"year band {y}")
    for j, lab in enumerate(("UZ", "Ptn", "Shown")):
        ok(ws.cell(row=HDR2, column=base + j).value == lab, f"sub-header {y} {lab}")
ok(ws.cell(row=HDR1, column=TOT).value == "All years", "totals band")

# ---- body ----------------------------------------------------------------
for n, row in enumerate(bc.itertuples()):
    r = FIRST + n
    ok(val(f"A{r}") == row.iso, f"A{r} iso out of order")
    ok(val(f"B{r}") == row.name, f"B{r} name")
    tot_uz = tot_ptn = tot_both = 0
    years_with = 0
    for i, y in enumerate(YEARS):
        base = 3 + 3 * i
        uz, ptn, both = want.get((row.iso, y), (0, 0, 0))
        gu = ws.cell(row=r, column=base).value
        gp = ws.cell(row=r, column=base + 1).value
        gb = ws.cell(row=r, column=base + 2).value
        ok(gu == uz, f"{row.iso} {y} UZ: {gu} != {uz}")
        ok(gp == ptn, f"{row.iso} {y} Ptn: {gp} != {ptn}")
        ok(gb == both, f"{row.iso} {y} Shown: {gb} != {both}")
        # Shown can never exceed either book: it is their intersection
        ok(gb <= gu and gb <= gp, f"{row.iso} {y} Shown exceeds a book")
        tot_uz += uz; tot_ptn += ptn; tot_both += both
        years_with += 1 if both > 0 else 0
    # the three All-years formulas must name every year column, once each
    for k, off in enumerate((0, 1, 2)):
        refs = ",".join(f"{get_column_letter(3 + 3 * i + off)}{r}" for i in range(len(YEARS)))
        ok(ws.cell(row=r, column=TOT + k).value == f"=SUM({refs})",
           f"{row.iso} All-years formula {k}")
    # and must agree with the per-year cells they sum
    ok(tot_both == row.cellyears_kept, f"{row.iso} Shown total != cellyears_kept")
    uzc, shc = get_column_letter(TOT), get_column_letter(TOT + 2)
    ok(ws.cell(row=r, column=PCT_COL).value == f'=IF({uzc}{r}=0,"",{shc}{r}/{uzc}{r})',
       f"{row.iso} percentage formula")
    ok(tot_uz == 0 or tot_both <= tot_uz, f"{row.iso} percentage would exceed 100%")
    shown_refs = ",".join(f"{get_column_letter(3 + 3 * i + 2)}{r}" for i in range(len(YEARS)))
    ok(ws.cell(row=r, column=YRS_COL).value == f'=COUNTIF({shown_refs},">0")',
       f"{row.iso} years-with formula")
    ok(0 <= years_with <= 8, f"{row.iso} years-with out of range")

# ---- total row -----------------------------------------------------------
TROW = LAST + 1
for c in range(3, TOT + 3):
    L = get_column_letter(c)
    ok(ws.cell(row=TROW, column=c).value == f"=SUM({L}{FIRST}:{L}{LAST})",
       f"total row column {L}")
grand_shown = sum(want.get((iso, y), (0, 0, 0))[2] for iso in bc["iso"] for y in YEARS)
ok(grand_shown == int(bc["cellyears_kept"].sum()),
   f"grand total shown {grand_shown} != {bc['cellyears_kept'].sum()}")
ok(ws.cell(row=TROW, column=2).value == "TOTAL - all partners", "total row label")

# nothing may sit below the notes, and the notes must be present
ok(isinstance(ws.cell(row=TROW + 2, column=1).value, str), "source note missing")
ok("cells matched" in (ws.cell(row=TROW + 3, column=1).value or ""), "verification note missing")

print(f"{checks:,} checks, {len(fails)} failed")
for f in fails[:25]:
    print("  FAIL", f)
raise SystemExit(1 if fails else 0)
